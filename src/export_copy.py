# src/alignment.py
"""
Экспорт и импорт для ручного выравнивания предложений.

Рабочий лист — «Английские»: каждая строка — EN-предложение,
колонка ru_№ указывает номер строки из листа «Русские».
Несколько EN-строк с одинаковым ru_№ = одно RU разбито на части.

Запуск:
    python src/alignment.py --list
    python src/alignment.py --export --ru 1 --en 1
    python src/alignment.py --import --file results/alignment.xlsx
"""

import sys
import sqlite3
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from config import DB_PATH, ALIGNMENT_FILE

# ── Цвета для Excel ───────────────────────────────────────────────────────────
FILL_AUTO      = PatternFill("solid", fgColor="C6EFCE")
FILL_SUGGESTED = PatternFill("solid", fgColor="FFEB9C")
FILL_MANUAL    = PatternFill("solid", fgColor="DDEBF7")
FILL_MANUAL_LO = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER    = PatternFill("solid", fgColor="2F5496")
FONT_HEADER    = Font(color="FFFFFF", bold=True)

SIM_LOW_THRESHOLD = 0.70

# ── утилиты ────────────────────────────────────────────────
def get_data(orig_text_id, translation_id):
    """Загружает данные из БД"""
    conn = sqlite3.connect(DB_PATH)
    
    orig = pd.read_sql(f"""
        SELECT sentence_id AS ru_id, position, sentence,
            search_word AS token_text, concept_phrase, gram_structure
        FROM sentences
        WHERE source_type='original' AND text_id={orig_text_id} AND language='ru'
        ORDER BY position
    """, conn)
    
    translation = pd.read_sql(f"""
        SELECT sentence_id AS en_id, position, sentence,
            search_word AS token_text, concept_phrase, gram_structure
        FROM sentences
        WHERE source_type='translation' AND translation_id={translation_id} AND language='en'
        ORDER BY position
    """, conn)
    
    align = pd.read_sql(f"""
        SELECT a.sentence_ru_id AS ru_id,
            a.sentence_en_id AS en_id,
            a.cosine_sim,
            COALESCE(a.auto_aligned, 0) AS auto_aligned
        FROM alignment a
        JOIN sentences sr ON sr.sentence_id = a.sentence_ru_id
        JOIN sentences se ON se.sentence_id = a.sentence_en_id
        WHERE sr.text_id = {orig_text_id} AND se.translation_id = {translation_id}
    """, conn)
    
    conn.close()
    
    # Нумеруем
    orig.insert(0, '№', range(1, len(orig) + 1))
    translation.insert(0, '№', range(1, len(translation) + 1))

    # ru_id_to_num = dict(zip(ru['sentence_id'], ru['ru_num']))
    # ru_num_to_text = dict(zip(ru['ru_num'], ru['sentence']))
    # ru_num_to_concept = dict(zip(ru['ru_num'], ru['concept_phrase']))
    # ru_num_to_grammar = dict(zip(ru['ru_num'], ru['gram_structure']))
    
    return orig, translation, align

# ── Translation shift ─────────────────────────────────────────────────────────
def translation_shift(cosine_sim) -> str:
    if pd.isna(cosine_sim):
        return ''
    sim = float(cosine_sim)
    if sim >= 0.85:
        return 'equivalent'
    if sim >= 0.65:
        return 'shift'
    return 'significant_shift'

# ── Список текстов ────────────────────────────────────────────────────────────
def get_texts_list(conn):
    print("\n📚 ДОСТУПНЫЕ ТЕКСТЫ")
    print("=" * 60)
    df_ru = pd.read_sql_query(
        "SELECT text_id, title, author FROM texts WHERE language='ru' ORDER BY title", conn)
    print("\n📖 Русские оригиналы (text_id):")
    for _, row in df_ru.iterrows():
        print(f"   {row['text_id']}: {row['title']} — {row['author']}")
    df_en = pd.read_sql_query(
        "SELECT translation_id, title, translator FROM translations WHERE language='en' ORDER BY title", conn)
    print("\n🌐 Английские переводы (translation_id):")
    for _, row in df_en.iterrows():
        print(f"   {row['translation_id']}: {row['title']} — {row['translator']}")
    return df_ru, df_en


# ── Экспорт ───────────────────────────────────────────────────────────────────

def export_for_alignment(conn, orig_id, translation_id, output_path=ALIGNMENT_FILE):
    # проверка наличия текста. Возможно не нужна
    ru_title = pd.read_sql_query(
        "SELECT title FROM texts WHERE text_id=? AND language='ru'",
        conn, params=(orig_id,))
    if ru_title.empty:
        print(f"❌ Русский текст с ID={orig_id} не найден")
        return None

    en_title = pd.read_sql_query(
        "SELECT title FROM translations WHERE translation_id=? AND language='en'",
        conn, params=(translation_id,))
    if en_title.empty:
        print(f"❌ Английский перевод с ID={translation_id} не найден")
        return None
    
    orig, translation, align = get_data(orig_id, translation_id)
    
    print(f"\n📖 Экспорт:")
    print(f"   RU: {ru_title.iloc[0]['title']} (text_id={orig_id})")
    print(f"   EN: {en_title.iloc[0]['title']} (translation_id={translation_id})")


    # ── 1.1. Дозаполнение cosine_sim для пар без него ────────────────────────
    missing = align[align['cosine_sim'].isna()]
    if not missing.empty:
        print(f"   🔢 Вычисляю cosine_sim для {len(missing)} пар без оценки...")
        try:
            import numpy as np
            from align_auto import get_model, _cosine_matrix
            model = get_model()
            cursor = conn.cursor()
            for _, row in missing.iterrows():
                ru_row = pd.read_sql_query(
                    "SELECT sentence FROM sentences WHERE sentence_id=?",
                    conn, params=(int(row['ru_id']),))
                en_row = pd.read_sql_query(
                    "SELECT sentence FROM sentences WHERE sentence_id=?",
                    conn, params=(int(row['en_id']),))
                if ru_row.empty or en_row.empty:
                    continue
                ru_vec = model.encode([ru_row.iloc[0]['sentence']], convert_to_numpy=True)
                en_vec = model.encode([en_row.iloc[0]['sentence']], convert_to_numpy=True)
                sim = float(_cosine_matrix(ru_vec, en_vec)[0, 0])
                cursor.execute(
                    "UPDATE alignment SET cosine_sim=? WHERE sentence_ru_id=? AND sentence_en_id=?",
                    (round(sim, 4), int(row['ru_id']), int(row['en_id'])))
                
                # перезаписываем датафрейм не залезая снова в бд
                align.loc[align['ru_id'] == row['ru_id'], 'cosine_sim'] = sim
            conn.commit()
            print(f"   ✅ cosine_sim обновлён")
        except Exception as e:
            print(f"   ⚠️  Не удалось вычислить sim: {e}")

    # ── 2. лист  ───────────────────────────────────
    df_all = (
        translation
        .merge(align, on='en_id', how='outer')                              # сохраняем все EN + все связи
        .merge(orig[['ru_id', 'sentence']], on='ru_id', how='outer', suffixes=('_en', '_ru'))  # сохраняем все RU
    )
    df_all['авто'] = df_all['auto_aligned'].map(
        lambda x: '✓' if x == 1 else ('?' if x == 2 else '')
    )
    # Выбираем колонки в нужном порядке
    en_cols = [
        'en_id', 'sentence_en', 
        'token_text', 
        'ru_id','sentence_ru', 
        'авто',
        'cosine_sim', 
        'auto_aligned'
    ]

    # ── 3. Запись в Excel ─────────────────────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_all[en_cols].to_excel(writer, sheet_name='Выравнивание', index=False)
    
    _format_workbook(output_path, df_all)

    print(f"\n✅ Сохранено: {output_path}")
    print(f"\n📝 Инструкция:")
    print(f"   1. Лист «Английские» — рабочий. Заполните ru_№ для пустых строк.")
    print(f"   2. 🟢 Зелёные — 1-й проход. 🟡 Жёлтые — 2-й проход (проверьте).")
    print(f"   3. translation_shift: equivalent ≥0.85 | shift ≥0.65 | significant_shift <0.65")
    print(f"   4. тональность — автоматически, поправьте при необходимости.")
    print(f"   5. Импорт: python src/alignment.py --import --file {output_path}")

    return output_path


# ── Форматирование Excel ──────────────────────────────────────────────────────
def _format_workbook(path, df_en):
    wb = load_workbook(path)

    col_widths = {
        # №  id  pos  sentence  token  concept  gram  тон  кол-во  EN№
        'Русские':    [5, 8, 8, 70, 14, 35, 45, 12, 10, 10],
        # №  id  pos  sentence  token  concept  тон  ru№  sim  авто  shift
        'Английские': [5, 8, 8, 70, 14, 35, 12, 8, 8, 6, 18],
        'Выравнивание': [8, 70, 14, 8, 70, 8, 14],  # ← добавили!
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = FONT_HEADER
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for i, width in enumerate(col_widths.get(sheet_name, []), 1):
            ws.column_dimensions[ws.cell(1, i).column_letter].width = width

        ws.freeze_panes = 'C2'

        if sheet_name in ['Английские', 'Выравнивание']:
            avto_col = df_en['авто'].fillna('')
            ru_col   = df_en['ru_id'].fillna('')
            sim_col  = df_en['cosine_sim'].fillna('')

            auto_rows      = set(i + 2 for i, v in enumerate(avto_col) if v == '✓')
            suggested_rows = set(i + 2 for i, v in enumerate(avto_col) if v == '?')
            manual_lo_rows = set()
            manual_hi_rows = set()
            for i, (ru_id, avto, sim_val) in enumerate(zip(ru_col, avto_col, sim_col)):
                # Проверяем, что ru_id не пустой И авто пустое
                if ru_id != '' and avto == '':
                    try:
                        sim_f = float(sim_val)
                        if sim_f < SIM_LOW_THRESHOLD:
                            manual_lo_rows.add(i + 2)
                        else:
                            manual_hi_rows.add(i + 2)
                    except (ValueError, TypeError):
                        manual_hi_rows.add(i + 2)

            for row in ws.iter_rows(min_row=2):
                r = row[0].row
                if r in auto_rows:
                    fill = FILL_AUTO
                elif r in suggested_rows:
                    fill = FILL_SUGGESTED
                elif r in manual_lo_rows:
                    fill = FILL_MANUAL_LO
                elif r in manual_hi_rows:
                    fill = FILL_MANUAL
                else:
                    continue
                for cell in row:
                    cell.fill = fill

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

    wb.save(path)


# ── Импорт ────────────────────────────────────────────────────────────────────
def import_alignment_from_excel(conn, excel_path=ALIGNMENT_FILE):
    if not Path(excel_path).exists():
        print(f"❌ Файл не найден: {excel_path}")
        return None

    df_en = pd.read_excel(excel_path, sheet_name='Английские', engine='openpyxl')
    df_ru = pd.read_excel(excel_path, sheet_name='Русские',    engine='openpyxl')

    ru_num_to_id = dict(zip(df_ru['№'].astype(int), df_ru['ru_id'].astype(int)))

    # Строки где авто=0 — удалить связь; остальные заполненные — добавить
    df_active = df_en[
        df_en['ru_№'].notna() &
        (df_en['ru_№'].astype(str).str.strip() != '')
    ].copy()

    df_delete = df_en[
        df_en['авто'].astype(str).str.strip() == '0'
    ].copy()

    print(f"\n📥 Импорт: {len(df_active)} заполненных EN-строк, удалений: {len(df_delete)}")

    cursor = conn.cursor()
    stats = {'inserted': 0, 'already': 0, 'deleted': 0, 'errors': 0}

    # Удаление помеченных строк (авто=0)
    for _, row in df_delete.iterrows():
        try:
            en_id = int(row['en_id'])
            cursor.execute(
                "DELETE FROM alignment WHERE sentence_en_id=?", (en_id,))
            deleted = cursor.rowcount
            if deleted:
                stats['deleted'] += deleted
                print(f"   🗑️  Удалена связь для en_id={en_id} ({deleted} запись)")
        except Exception as e:
            print(f"   ⚠️  Ошибка удаления en_id={row.get('en_id')}: {e}")
            stats['errors'] += 1

    # Добавление / проверка остальных пар
    for _, row in df_active.iterrows():
        try:
            en_id  = int(row['en_id'])
            # Пропустить строки помеченные на удаление
            if str(row.get('авто', '')).strip() == '0':
                continue
            ru_num = int(float(str(row['ru_№']).strip()))
            ru_id  = ru_num_to_id.get(ru_num)

            if ru_id is None:
                print(f"   ⚠️  ru_№={ru_num} не найден в листе «Русские»")
                stats['errors'] += 1
                continue

            cursor.execute(
                "SELECT alignment_id FROM alignment WHERE sentence_ru_id=? AND sentence_en_id=?",
                (ru_id, en_id))

            if cursor.fetchone():
                stats['already'] += 1
            else:
                cursor.execute(
                    "INSERT INTO alignment (sentence_ru_id, sentence_en_id, auto_aligned) VALUES (?,?,0)",
                    (ru_id, en_id))
                stats['inserted'] += 1

        except Exception as e:
            print(f"   ⚠️  Ошибка строки en_id={row.get('en_id')}: {e}")
            stats['errors'] += 1

    conn.commit()

    print(f"\n✅ Импорт завершён:")
    print(f"   Добавлено: {stats['inserted']}")
    print(f"   Уже было:  {stats['already']}")
    print(f"   Удалено:   {stats['deleted']}")
    print(f"   Ошибок:    {stats['errors']}")

    return stats


def main():
    import argparse
    parser = argparse.ArgumentParser(description='Выравнивание предложений RU↔EN')
    parser.add_argument('--list',   action='store_true')
    parser.add_argument('--export', action='store_true')
    parser.add_argument('--import', dest='import_mode', action='store_true')
    parser.add_argument('--ru',   type=int)
    parser.add_argument('--en',   type=int)
    parser.add_argument('--file', type=str, default=str(ALIGNMENT_FILE))
    args = parser.parse_args()

    conn = sqlite3.connect(DB_PATH)

    if args.list:
        get_texts_list(conn)
    elif args.export:
        if not args.ru or not args.en:
            print("❌ Укажите --ru и --en (см. --list)")
        else:
            export_for_alignment(conn, args.ru, args.en, args.file)
    elif args.import_mode:
        import_alignment_from_excel(conn, args.file)
    else:
        parser.print_help()

    conn.close()


if __name__ == "__main__":
    main()
